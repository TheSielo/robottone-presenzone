import logging
import schedule
import time

from datetime import datetime
from telegram import Update, Bot, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import CallbackContext, CommandHandler, Updater, CallbackQueryHandler, MessageHandler, Filters

#You have to create 'tokens.py' in the same folder as this file, and there declare
#two strings, 'discordToken' and 'telegramToken', and assign them the correct values.
#The file is not commited to avoid uploading the tokens to the git repository.
from tokens import token
from data import runningOperations
from registration import register, continueRegistration, deleteUser
from stUtils import getId
import preferences

from openpyxl import load_workbook

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

telegramBot: Bot = Bot(token)
updater = Updater(token=token, use_context=True)  

TYPE_NO_REMAINING = 0
TYPE_HOLIDAY = 1
TYPE_ILLNESS = 2

START_EXPLANATION = '''Drin Dron Robotton:\n\n
Prima di tutto, esegui la registrazione con il comando /register, altrimenti non riuscirai a fare un piffero di niente!\n\n
/start - Mostra le info che stai leggendo proprio ora.\n\n
/register - Registrati per inserire i tuoi dati e iniziare a utilizzare il Robottone Presenzone.\n\n
/delete - Elimina i tuoi dati e cancella la tua utenza.\n\n
/insert - Inserisci manualmente gli orari di oggi senza aspettare l'intervento del Robottone Presenzone.\n\n
/edit - Modifica le ore per un giorno specifico selezionato.\n\n
/sendme - Scarica il Foglione Presenzone senza attendere che ti venga inviato dal Robottone.\n\n
'''

#Check that makes sense to continue executing the command.
def checkArguments(update, context, hasArgs):
    #This check MUST be used before answering any command to avoid other groups to use this bot
    '''if not str(update.effective_chat.id) == ids.GROUP_ID and not str(update.effective_chat.id) == ids.TEST_GROUP_ID:
        context.bot.send_message(chat_id=update.effective_chat.id, text="Remove this bot from this chat, or prepare to pay for your actions.")
        print('WRONG CHAT USAGE ATTEMPT!!!')
        return False

    #If the command needs some arguments but there's none, inform the user and return False to abort the execution of the command.
    if hasArgs and len(context.args) == 0:
        context.bot.send_message(chat_id=update.effective_chat.id, text='Si ma dammi sti parametri pero\'')
        return False'''

    return True


#Show handy instruction about using this bot.
def start(update: Update, context: CallbackContext):
    if not checkArguments(update, context, False):
        return

    update.message.reply_text(START_EXPLANATION)

def insert(update: Update, context: CallbackContext):
    userId = getId(update)
    preferences.setMode(userId, preferences.MODE_INSERTING)
    insertTime(update, context)

#Add time for today
def insertTime(update: Update, context: CallbackContext, userId: str = None):
    if update:
        id = getId(update)
    else:
        id = userId
    
    if checkState(context, id):
        keyboard = [[InlineKeyboardButton('0', callback_data='h:0')],
                [InlineKeyboardButton('1', callback_data='h:1'),
                InlineKeyboardButton('2', callback_data='h:2'),
                InlineKeyboardButton('3', callback_data='h:3'),
                InlineKeyboardButton('4', callback_data='h:4')],

                [InlineKeyboardButton('5', callback_data='h:5'),
                InlineKeyboardButton('6', callback_data='h:6'),
                InlineKeyboardButton('7', callback_data='h:7'),
                InlineKeyboardButton('8', callback_data='h:8')],
                [InlineKeyboardButton('Non compilare questo giorno', callback_data='d:1')]]

        reply_markup = InlineKeyboardMarkup(keyboard)
        mode = preferences.getMode(id)
        if mode == preferences.MODE_EDITING:
            text = 'Quante ore hai lavorato quel giorno?'
        else:
            text = 'Quante ore hai lavorato oggi?'
        updater.bot.send_message(chat_id=id, text=text, reply_markup=reply_markup)

def editTime(update: Update, context: CallbackContext):
    userId = getId(update)
    if checkState(context, userId):
        preferences.setMode(userId, preferences.MODE_EDITING)
        keyboard = [
                [InlineKeyboardButton('1', callback_data='e:1'),
                InlineKeyboardButton('2', callback_data='e:2'),
                InlineKeyboardButton('3', callback_data='e:3'),
                InlineKeyboardButton('4', callback_data='e:4')],
                [InlineKeyboardButton('5', callback_data='e:5'),
                InlineKeyboardButton('6', callback_data='e:6'),
                InlineKeyboardButton('7', callback_data='e:7'),
                InlineKeyboardButton('8', callback_data='e:8')],
                [InlineKeyboardButton('9', callback_data='e:9'),
                InlineKeyboardButton('10', callback_data='e:10'),
                InlineKeyboardButton('11', callback_data='e:11'),
                InlineKeyboardButton('12', callback_data='e:12')],
                [InlineKeyboardButton('13', callback_data='e:13'),
                InlineKeyboardButton('14', callback_data='e:14'),
                InlineKeyboardButton('15', callback_data='e:15'),
                InlineKeyboardButton('16', callback_data='e:16')],
                [InlineKeyboardButton('17', callback_data='e:17'),
                InlineKeyboardButton('18', callback_data='e:18'),
                InlineKeyboardButton('19', callback_data='e:19'),
                InlineKeyboardButton('20', callback_data='e:20')],
                [InlineKeyboardButton('21', callback_data='e:21'),
                InlineKeyboardButton('22', callback_data='e:22'),
                InlineKeyboardButton('23', callback_data='e:23'),
                InlineKeyboardButton('24', callback_data='e:24')],
                [InlineKeyboardButton('25', callback_data='e:25'),
                InlineKeyboardButton('26', callback_data='e:26'),
                InlineKeyboardButton('27', callback_data='e:27'),
                InlineKeyboardButton('28', callback_data='e:28')],
                [InlineKeyboardButton('29', callback_data='e:29'),
                InlineKeyboardButton('30', callback_data='e:30'),
                InlineKeyboardButton('31', callback_data='e:31')]]

        reply_markup = InlineKeyboardMarkup(keyboard)
        message_reply_text = 'Quale giorno vuoi modificare?'
        context.bot.send_message(chat_id=userId, text=message_reply_text, reply_markup=reply_markup)
    

#Save the selected number of hours and ask for the minutes
def setMinutes(update: Update, context: CallbackContext):
    keyboard = [[InlineKeyboardButton('00', callback_data='m:00'),
            InlineKeyboardButton('15', callback_data='m:15'),
            InlineKeyboardButton('30', callback_data='m:30'),
            InlineKeyboardButton('45', callback_data='m:45')]]


    reply_markup = InlineKeyboardMarkup(keyboard)
    message_reply_text = 'E quanti minuti?'
    context.bot.send_message(chat_id=update.effective_chat.id, text=message_reply_text, reply_markup=reply_markup)

def setType(update: Update, context: CallbackContext):
    keyboard = [[InlineKeyboardButton('Ferie', callback_data='t:1'),
            InlineKeyboardButton('Malattia', callback_data='t:2')]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    message_reply_text = 'Come segno le ore rimanenti?'
    context.bot.send_message(chat_id=getId(update), text=message_reply_text, reply_markup=reply_markup)


def buttonPressed(update: Update, context: CallbackContext):
    userId = getId(update)
    context.bot.delete_message(chat_id=userId, message_id=update.effective_message.message_id)
    
    parts: list[str] = update.callback_query.data.split(':')
    command: str = parts[0]
    value = int(parts[1])

    if command == 'h':
        day = 0
        if userId in runningOperations and 'e' in runningOperations[userId]:
            day = runningOperations[userId]['e']
        runningOperations[userId] = {'h':value, 'e':day}
        setMinutes(update, context)
    elif command == 'm':
        runningOperations[userId]['m'] = value
        hours = runningOperations[userId]['h']
        if int(hours) >= 8:
            runningOperations[userId]['t'] = TYPE_NO_REMAINING
            day = runningOperations[userId]['e']
            writeToday(update, context, day)
        else:
            setType(update, context)
    elif command == 't':
        runningOperations[userId]['t'] = value
        day = runningOperations[userId]['e']
        writeToday(update, context, day)
    elif command == 'e':
        runningOperations[userId] = {'e':int(value)}
        insertTime(update, context)
    elif command == 'd':
        preferences.setMode(userId, preferences.MODE_DELETING)
        day = 0
        if userId in runningOperations and 'e' in runningOperations[userId]:
            day = runningOperations[userId]['e']
        writeToday(update, context, day)


def writeToday(update: Update, context: CallbackContext, day: int):
    userId = getId(update)
    mode = preferences.getMode(userId)

    date = datetime.today()
    filename = preferences.getUserFile(userId)
    wb = load_workbook(filename=filename)
    ws = wb.worksheets[date.month-1]

    if day == 0:
        dayToWrite = date.day
    else:
        dayToWrite = day

    ws['c4'] = preferences.loadConfig(userId, preferences.KEY_USER_NAME)
    ws['c6'] = preferences.loadConfig(userId, preferences.KEY_MANAGER_NAME)
    
    if mode == preferences.MODE_DELETING:
        ws.cell(row=dayToWrite+9, column=3).value = ''
        ws.cell(row=dayToWrite+9, column=4).value = ''
        ws.cell(row=dayToWrite+9, column=5).value = ''
    else:
        hours = runningOperations[userId]['h']
        minutes = runningOperations[userId]['m']
        type = runningOperations[userId]['t']

        time = ('%s%s' % (hours, convertMinutes(minutes)))
        if time != '0':
            ws.cell(row=dayToWrite+9, column=3).value = time

        if type != TYPE_NO_REMAINING:
            if minutes == 0:
                remainingHours = 8 - hours
                remainingMinutes = 0
            else:
                remainingHours = 7 - hours
                remainingMinutes = 60 - minutes

            if type == TYPE_HOLIDAY:
                remainingCol = 4
                emptyCol = 5
            elif type == TYPE_ILLNESS:
                remainingCol = 5
                emptyCol = 4

            ws.cell(row=dayToWrite+9, column=remainingCol).value = ('%d%s' % (remainingHours, convertMinutes(remainingMinutes)))
            ws.cell(row=dayToWrite+9, column=emptyCol).value = ''
        else:
            ws.cell(row=dayToWrite+9, column=4).value = ''
            ws.cell(row=dayToWrite+9, column=5).value = ''
    
    wb.save(filename)
    runningOperations.pop(userId, None)
    preferences.setMode(userId, preferences.MODE_NOTHING)

    weekend = [4,5,6]
    if mode == preferences.MODE_EDITING:
        text = "Grazie!"
    elif date.weekday() in weekend:
        text = 'Grazie! A Lunedì! (Spero)'
    else:
        text = 'Grazie! A domani!'
        
    context.bot.send_message(chat_id=getId(update), text=text)

def isLastDayOfTheMonth(date: datetime):
    return False

    
def convertMinutes(minutes: int):
    if minutes == 0:
        return ''
    elif minutes == 15:
        return ',25'
    elif minutes == 30:
        return ',50'
    elif minutes == 45:
        return ',75'

def checkState(context: CallbackContext, userId: str):
    state = preferences.loadConfig(userId, preferences.KEY_STATE)
    if state != None and int(state) >= preferences.STATE_REGISTERED:
        return True
    else:
        text = "Prima di continuare effettua la registrazione con il comando /register!"
        context.bot.send_message(chat_id=userId, text=text)
        return False


def sendSheet(update: Update, context: CallbackContext, userId: str = None):
    if update:
        user = getId(update)
    else:
        user = userId

    if checkState(context, user):
        document = open(preferences.getUserFile(user), 'rb')
        updater.bot.send_document(chat_id=user, document=document, filename='Foglio presenze.xlsx')



def forwarder(update: Update, context: CallbackContext):
    userId = getId(update)
    state = preferences.loadConfig(userId, preferences.KEY_STATE)
    if state != None and int(state) < preferences.STATE_REGISTERED:
        continueRegistration(update, context, int(state))
    else:
        text = "Usa uno dei comandi disponibili! Se non ti sei ancora registrato, inizia la registrazione con il comando /register!"
        context.bot.send_message(chat_id=userId, text=text)

# Script start
dispatcher = updater.dispatcher

#Analyze all messages that are not commands
forwardHandler = MessageHandler(Filters.text & (~Filters.command), forwarder)
dispatcher.add_handler(forwardHandler)

# Buttons callback
dispatcher.add_handler(CallbackQueryHandler(buttonPressed))

# /start command
startHandler = CommandHandler('start', start)
dispatcher.add_handler(startHandler)

# /register command
registerHandler = CommandHandler('register', register)
dispatcher.add_handler(registerHandler)

# /delete command
deleteHandler = CommandHandler('delete', deleteUser)
dispatcher.add_handler(deleteHandler)

# /insert command
insertHandler = CommandHandler('insert', insert)
dispatcher.add_handler(insertHandler)

# /edit command
editHandler = CommandHandler('edit', editTime)
dispatcher.add_handler(editHandler)

# /send sheet command
sendSheetHandler = CommandHandler('sendme', sendSheet)
dispatcher.add_handler(sendSheetHandler)

# Start Telegram bot
updater.start_polling()

def sendQuestions():
    users = preferences.getSections()
    for user in users:
       insertTime(None, None, userId=user) 


TIME = '19:00'
schedule.every().monday.at(TIME).do(sendQuestions)
schedule.every().tuesday.at(TIME).do(sendQuestions)
schedule.every().wednesday.at(TIME).do(sendQuestions)
schedule.every().thursday.at(TIME).do(sendQuestions)
schedule.every().friday.at(TIME).do(sendQuestions)

while True:
    schedule.run_pending()
    time.sleep(30)
