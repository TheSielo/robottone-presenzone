from datetime import datetime
from telegram import Update
from telegram.ext import CallbackContext

from data import runningOperations
from stUtils import getId
from preferences import getMode, getUserFile, loadConfig, setMode
from preferences import KEY_USER_NAME, KEY_MANAGER_NAME, MODE_DELETING, MODE_EDITING, MODE_NOTHING

from openpyxl import load_workbook

TYPE_NO_REMAINING = 0
TYPE_HOLIDAY = 1
TYPE_ILLNESS = 2

def writeToday(update: Update, context: CallbackContext, day: int):
    userId = getId(update)
    mode = getMode(userId)

    date = datetime.today()
    filename = getUserFile(userId)
    wb = load_workbook(filename=filename)
    ws = wb.worksheets[date.month-1]

    if day == 0:
        dayToWrite = date.day
    else:
        dayToWrite = day

    ws['c4'] = loadConfig(userId, KEY_USER_NAME)
    ws['c6'] = loadConfig(userId, KEY_MANAGER_NAME)
    
    if mode == MODE_DELETING:
        ws.cell(row=dayToWrite+9, column=3).value = ''
        ws.cell(row=dayToWrite+9, column=4).value = ''
        ws.cell(row=dayToWrite+9, column=5).value = ''
    else:
        hours = runningOperations[userId]['h']
        minutes = runningOperations[userId]['m']
        type = runningOperations[userId]['t']

        time = ('%s%s' % (hours, convertMinutes(minutes)))
        if time != '0':
            ws.cell(row=dayToWrite+9, column=3).value = time

        if type != TYPE_NO_REMAINING:
            if minutes == 0:
                remainingHours = 8 - hours
                remainingMinutes = 0
            else:
                remainingHours = 7 - hours
                remainingMinutes = 60 - minutes

            if type == TYPE_HOLIDAY:
                remainingCol = 4
                emptyCol = 5
            elif type == TYPE_ILLNESS:
                remainingCol = 5
                emptyCol = 4

            ws.cell(row=dayToWrite+9, column=remainingCol).value = ('%d%s' % (remainingHours, convertMinutes(remainingMinutes)))
            ws.cell(row=dayToWrite+9, column=emptyCol).value = ''
        else:
            ws.cell(row=dayToWrite+9, column=4).value = ''
            ws.cell(row=dayToWrite+9, column=5).value = ''
    
    wb.save(filename)
    runningOperations.pop(userId, None)
    setMode(userId, MODE_NOTHING)

    weekend = [4,5,6]
    if mode == MODE_EDITING:
        text = "Grazie!"
    elif date.weekday() in weekend:
        text = 'Grazie! A Lunedì! (Spero)'
    else:
        text = 'Grazie! A domani!'
        
    context.bot.send_message(chat_id=getId(update), text=text)


def isLastDayOfTheMonth(date: datetime):
    return False

    
def convertMinutes(minutes: int):
    if minutes == 0:
        return ''
    elif minutes == 15:
        return ',25'
    elif minutes == 30:
        return ',50'
    elif minutes == 45:
        return ',75'
